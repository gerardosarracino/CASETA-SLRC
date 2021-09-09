# -*- coding: utf-8 -*-
import json
import logging
import werkzeug.utils
from odoo.exceptions import UserError, ValidationError
from odoo import http, _
from odoo.http import request
from odoo.api import call_kw, Environment
from odoo.osv.expression import AND
from odoo.tools import convert
import calendar
from datetime import datetime, date, time, timedelta
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl.styles import colors
from openpyxl.styles import Font, Color, NamedStyle
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

import time
from datetime import datetime

_logger = logging.getLogger(__name__)

import odoorpc


class MandarVenta(http.Controller):
    @http.route('/venta_controller/<id_usuario>/<func_id>', type='http', auth='public', methods=["GET","POST"], csrf=False, cors='*')
    def index(self, id_usuario, func_id): # func_id
        try:
            search_usuario = http.request.env['res.users'].sudo().search([('id', '=', int(id_usuario))])  # ('id', '=', int(id_usuario))

            orden = http.request.env['pos.order']
            # carril = ''
            # sesion_actual = ''
            # fecha_hoy = datetime.now()
            for user in search_usuario:
                browse_usuario = http.request.env['res.users'].browse(user)
                usuario = browse_usuario.id
                search_sesion = http.request.env['pos.session'].sudo().search([('state', '=', 'opened'),('user_id.id', '=', int(browse_usuario.id))])  # ('id', '=', int(id_usuario))
                for sesion in search_sesion:
                    sesion_actual = sesion.id
                    # carril = sesion.config_id

                    datos = {
                        # 'name': str(carril)+'/Residente',
                        'session_id': sesion_actual,
                        'pos_reference': func_id, # ID DEL TICKET
                        # 'cashier': usuario,
                        'user_id': user.id,
                        'amount_total': 0,
                        'state': 'paid',
                        'amount_tax': 0,
                        'amount_paid': 0,
                        'amount_return': 0,
                        'company_id': 1,
                        'payment_ids': [],
                        'pricelist_id': 1,
                    }
                    order = orden.create(datos)

                    orden_creada = http.request.env['pos.order'].sudo().search([('id', '=', order.id)])
                    search_producto = http.request.env['product.product'].sudo().search([('name', '=', 'RESIDENTE')])
                    producto = ''
                    taxes = ''
                    for pro in search_producto:
                        # browse_producto = http.request.env['product.product'].browse(pro)
                        producto = pro.id
                        taxes = pro.taxes_id.id

                    tabla_productos = {'lines': [[0, 0, {
                        'product_id': producto,
                        'order_id': order.id, # ID DE LA ORDEN
                        'qty': 1,
                        'discount': 0,
                        'tax_ids_after_fiscal_position': taxes,
                        'price_subtotal': 0,
                        'price_subtotal_incl': 0,
                    }]]}
                    agregar_tabla = orden_creada.write(tabla_productos)

                    payment = http.request.env['pos.payment']
                    tabla_pago = {
                        'pos_order_id': orden_creada.id,
                        'session_id': sesion_actual, # id de la sesion
                        'payment_method_id': 1,
                        'amount': 0,
                    }
                    agregar_tablax2 = payment.create(tabla_pago)

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


class MandarVentaCompleta(http.Controller):
    @http.route('/venta_pos_controller/<id_usuario>/<func_id>/<monto_siva>/<monto_iva>/<id_sesion>/<id_producto>', type='http', auth='public', methods=["GET","POST"], csrf=False, cors='*')
    def index(self, id_usuario, func_id, monto_siva, monto_iva, id_sesion, id_producto): # func_id
        try:
            search_usuario = http.request.env['res.users'].sudo().search([('name', '=', str(id_usuario))])  # ('id', '=', int(id_usuario))
            print('ENTRO CONTROLLER', id_sesion)
            orden = http.request.env['pos.order']
            # carril = ''
            # sesion_actual = ''
            # fecha_hoy = datetime.now()
            for user in search_usuario:
                browse_usuario = http.request.env['res.users'].browse(user)
                usuario = browse_usuario.id
                search_sesion = http.request.env['pos.session'].sudo().search([('state', '=', 'opened'),('id', '=', int(id_sesion))])  # ('id', '=', int(id_usuario))
                for sesion in search_sesion:
                    sesion_actual = sesion.id
                    # carril = sesion.config_id

                    datos = {
                        # 'name': str(carril)+'/Residente',
                        'session_id': sesion_actual,
                        'pos_reference': func_id, # ID DEL TICKET
                        # 'cashier': usuario,
                        'user_id': user.id,
                        'amount_total': float(monto_siva) + float(monto_iva), # MONTO S/IVA
                        'state': 'paid',
                        'amount_tax': float(monto_iva), # IVA
                        'amount_paid': float(monto_siva) + float(monto_iva), # MONTO C/IVA
                        'amount_return': 0,
                        'company_id': 1,
                        'payment_ids': [],
                        'pricelist_id': 1,
                    }
                    print(datos)
                    order = orden.create(datos)

                    orden_creada = http.request.env['pos.order'].sudo().search([('id', '=', order.id)])
                    search_producto = http.request.env['product.product'].sudo().search([('name', '=', str(id_producto))])
                    producto = ''
                    taxes = ''
                    for pro in search_producto:
                        producto = pro.id
                        taxes = pro.taxes_id.id
                    print(taxes)
                    tabla_productos = {'lines': [[0, 0, {
                        'product_id': producto,
                        'order_id': order.id, # ID DE LA ORDEN
                        'qty': 1,
                        'discount': 0,
                        'tax_ids_after_fiscal_position': taxes,
                        'price_unit': float(monto_siva),
                        'price_subtotal': float(monto_siva),
                        'price_subtotal_incl': float(monto_siva) + float(monto_iva),
                    }]]}
                    print(tabla_productos)
                    agregar_tabla = orden_creada.write(tabla_productos)

                    payment = http.request.env['pos.payment']
                    tabla_pago = {
                        'pos_order_id': orden_creada.id,
                        'session_id': sesion_actual, # id de la sesion
                        'payment_method_id': 1,
                        'amount': float(monto_siva) + float(monto_iva),
                    }
                    agregar_tablax2 = payment.create(tabla_pago)

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


class CerrarSesionPos(http.Controller):
    @http.route('/cerrar_sesion_pos/<pos_session>/<id_usuario>', type='http', auth='public', csrf=False, cors='*')
    def index(self, pos_session, id_usuario):
        try:
            print(pos_session, id_usuario)
            search_usuario = http.request.env['res.users'].sudo().search(
                [('name', '=', str(id_usuario))])  # ('id', '=', int(id_usuario))
            nombre_usuario = ''
            cajero_id = ''
            for user in search_usuario:
                browse_usuario = http.request.env['res.users'].browse(user.id)
                nombre_usuario = browse_usuario.name
                cajero_id = browse_usuario.id
            print(nombre_usuario, cajero_id)
            if nombre_usuario == 'Administrator': # ES EL ADMINISTRADOR DEL CARRIL NO HACER CAMBIO
                pass
            else:
                search_sesion = http.request.env['pos.session'].sudo().search([('id', '=', int(pos_session))])  # ('id', '=', int(id_usuario))
                print('ENTRO CONTROLLER', search_sesion, search_usuario)
                carril = ''
                for ss in search_sesion:
                    browse_sesion = http.request.env['pos.session'].browse(ss.id)
                    print(int(cajero_id), int(browse_sesion.user_id.id), ' IF')
                    print(browse_sesion.state, 'ESTADO X1')
                    if int(cajero_id) == int(browse_sesion.user_id.id):
                        pass
                    else:
                        carril = browse_sesion.config_id.id
                        for session in browse_sesion:
                            for statement in session.statement_ids:
                                if (statement != session.cash_register_id) and (
                                        statement.balance_end != statement.balance_end_real):
                                    statement.write({'balance_end_real': statement.balance_end})
                        for session in browse_sesion:
                            if session.state == 'closed':
                                raise UserError(_('This session is already closed.'))
                            browse_sesion.write({'state': 'closing_control', 'stop_at': datetime.now()})
                            if not session.config_id.cash_control:
                                browse_sesion.action_pos_session_close()
                        print(browse_sesion.state, 'ESTADO')

                        print('Termino')
                if carril:
                    print(carril, 'cc', cajero_id)
                    sesion = http.request.env['pos.session']

                    datos = {
                        'config_id': carril,
                        'user_id': int(cajero_id),
                        'start_at': datetime.now(),
                        'state': 'opened',
                    }
                    r = sesion.create(datos)
                    print(datos)

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


class ReporteExcel(http.Controller):
    @http.route('/reporte_excel/reporte_excel/', auth='public')
    def index(self, id_informe):
        try:
            # workbook = load_workbook(filename="/home/gerardo/Developments/odoo13/extra-addons/slrc/static/reporte_excel/reporte_concentrado_aforo_ingrensos.xlsx")
            workbook = load_workbook(filename="/root/Developments/odoo13/extra-addons/slrc/static/reporte_excel/reporte_concentrado_aforo_ingrensos.xlsx") # /opt/SLRC/src/slrc/static/src/reporte_excel/reporte_concentrado_aforo_ingrensos.xlsx
            sheet = workbook.active
            wb = Workbook()

            wizard = http.request.env['informe.excel.wizard'].sudo().search([])
            modelo = http.request.env['pos.order']

            meses = ["Unknown",
                     "ENERO",
                     "FEBRERO",
                     "MARZO",
                     "ABRIL",
                     "MAYO",
                     "JUNIO",
                     "JULIO",
                     "AGOSTO",
                     "SEPTIEMBRE",
                     "OCTUBRE",
                     "NOVIEMBRE",
                     "DICIEMBRE"]

            for w in wizard:
                w_browse = http.request.env['informe.excel.wizard'].browse(int(id_informe))
                ventas = modelo.sudo().search([('date_order', '>', w_browse.start_date),
                                               ('date_order', '<', w_browse.end_date)], order='date_order asc')

                producto = http.request.env['product.template'].sudo().search([]) # BUSQUEDA DE LOS PRECIOS
                for proc in producto:
                    proc_browse = http.request.env['product.template'].browse(proc.id)
                    letra_columa_precio = get_column_letter(2)
                    if proc_browse.name == 'MOTOCICLETA':
                        sheet[letra_columa_precio + '7'] = proc_browse.list_price # PRECIO MOTOCICLETA
                    if proc_browse.name == 'AUTO':
                        sheet[letra_columa_precio + '16'] = proc_browse.list_price # PRECIO AUTO
                    if proc_browse.name == 'AUTO + 1 EJE':
                        sheet[letra_columa_precio + '19'] = proc_browse.list_price # PRECIO AUTO 1 EJE
                    if proc_browse.name == 'AUTO + 2 EJE':
                        sheet[letra_columa_precio + '22'] = proc_browse.list_price # PRECIO AUTO 2 EJES
                    if proc_browse.name == 'AUTOBUS':
                        sheet[letra_columa_precio + '32'] = proc_browse.list_price # PRECIO AUTOBUS
                    if proc_browse.name == 'CAMION 2 EJES':
                        sheet[letra_columa_precio + '42'] = proc_browse.list_price # PRECIO CAMION 2 EJES
                    if proc_browse.name == 'CAMION 5 EJES':
                        sheet[letra_columa_precio + '46'] = proc_browse.list_price # PRECIO CAMION 5 EJES
                    if proc_browse.name == 'CAMION + 7 EJES':
                        sheet[letra_columa_precio + '50'] = proc_browse.list_price # PRECIO CAMION 7+ EJES
                # ////////////////////// # CONCENTRADO DE AFORO E INGRESOS + ANIO
                letra_columa_ano = get_column_letter(1)
                fecha_concentrado_aforo = w_browse.end_date.year
                sheet[letra_columa_ano + '2'] = 'CONCENTRADO DE AFORO E INGRESOS ' + str(fecha_concentrado_aforo)
                # ////////////////////// # AFORO E INGRESOS DE + FECHA
                letra_columa_fecha = get_column_letter(8)
                fecha_aforo = meses[w_browse.start_date.month] + " A " + meses[w_browse.end_date.month] +\
                                     " DE " + str(w_browse.end_date.year)
                sheet[letra_columa_fecha + '2'] = 'AFORO E INGRESOS DE ' + str(fecha_aforo)
                # ////////////////////// # FECHA DE IMPRESION
                letra_columa_fechahoy = get_column_letter(13)
                ahora = datetime.now()
                sheet[letra_columa_fechahoy + '3'] = 'FECHA: ' + str(ahora)
                # //////////////////////
                cont_motos = 0
                cont_autos = 0 # auto normal
                cont_autos_1eje = 0 # auto 1 eje
                cont_autos_2eje = 0 # auto 2 ejes
                cont_autobuses = 0 # TARIFA B2-B4 (AUTOBUSES DE 2 A 4 EJES)
                cont_camion_2ejes = 0 # TARIFA C2-C4   (CAMIONES DE 2 A 4 EJES)
                cont_camion_4ejes = 0 # TARIFA C2-C4   (CAMIONES DE 2 A 4 EJES)
                cont_camion_5ejes = 0 # TARIFA C5-C6   (CAMIONES DE 2 A 4 EJES)
                cont_camion_6ejes = 0 # TARIFA C5-C6   (CAMIONES DE 2 A 4 EJES)
                cont_camion_7ejes = 0 # TARIFA C5-C6   (CAMIONES DE 2 A 4 EJES)
                cont_camion_s9ejes = 0 # TARIFA C5-C6   (CAMIONES DE 2 A 4 EJES)
                # //////////////////////
                mes_anterior = 0
                for i in ventas:
                    pos_browse = http.request.env['pos.order'].browse(i.id)
                    print(pos_browse.date_order)
                    mes = pos_browse.date_order.month

                    if mes_anterior != 0:
                        if pos_browse.date_order.month != mes_anterior:
                            cont_motos = 0
                            cont_autos = 0  # auto normal
                            cont_autos_1eje = 0
                            cont_autos_2eje = 0
                            cont_autobuses = 0  # TARIFA B2-B4 (AUTOBUSES DE 2 A 4 EJES)
                            cont_camion_2ejes = 0  # TARIFA C2-C4   (CAMIONES DE 2 A 4 EJES)
                            cont_camion_4ejes = 0  # TARIFA C2-C4   (CAMIONES DE 2 A 4 EJES)
                            cont_camion_5ejes = 0  # TARIFA C5-C6   (CAMIONES DE 5 A 6 EJES)
                            cont_camion_6ejes = 0  # TARIFA C5-C6   (CAMIONES DE 5 A 6 EJES)
                            cont_camion_7ejes = 0  # TARIFA C7-C9   (CAMIONES DE 7 A 9 EJES)
                            cont_camion_s9ejes = 0  # TARIFA C7-C9   (CAMIONES DE 7 A 9 EJES)

                    for lines in pos_browse.lines:
                        if lines.product_id.name == 'MOTOCICLETA':
                            cont_motos += 1
                        if lines.product_id.name == 'AUTO':
                            cont_autos += 1
                        if lines.product_id.name == 'AUTO + 1 EJE':
                            cont_autos_1eje += 1
                        if lines.product_id.name == 'AUTO + 2 EJE':
                            cont_autos_2eje += 1
                        if lines.product_id.name == 'AUTOBUS':
                            cont_autobuses += 1
                        if lines.product_id.name == 'CAMION 2 EJES':
                            cont_camion_2ejes += 1
                        if lines.product_id.name == 'CAMION 4 EJES':
                            cont_camion_4ejes += 1
                        if lines.product_id.name == 'CAMION 5 EJES':
                            cont_camion_5ejes += 1
                        if lines.product_id.name == 'CAMION 6 EJES':
                            cont_camion_6ejes += 1
                        if lines.product_id.name == 'CAMION + 7 EJES':
                            cont_camion_7ejes += 1

                    letra_columa = ""
                    if mes == 1:  # ENERO
                        letra_columa = get_column_letter(3) # COLUMNA C
                    elif mes == 2:  # FEBRERO
                        letra_columa = get_column_letter(4)  # COLUMNA D
                    elif mes == 3:  # MARZO
                        letra_columa = get_column_letter(5)  # COLUMNA E
                    elif mes == 4:  # ABRIL
                        letra_columa = get_column_letter(6)  # COLUMNA F
                    elif mes == 5:  # MAYO
                        letra_columa = get_column_letter(7)  # COLUMNA G
                    elif mes == 6:  # JUNIO
                        letra_columa = get_column_letter(8)  # COLUMNA H
                    elif mes == 7:  # JULIO
                        letra_columa = get_column_letter(9)  # COLUMNA I
                    elif mes == 8:  # AGOSTO
                        letra_columa = get_column_letter(10)  # COLUMNA J
                    elif mes == 9:  # SEPTIEMBRE
                        letra_columa = get_column_letter(11)  # COLUMNA K
                    elif mes == 10:  # OCTUBRE
                        letra_columa = get_column_letter(12)  # COLUMNA L
                    elif mes == 11:  # NOVIEMBRE
                        letra_columa = get_column_letter(13)  # COLUMNA M
                    elif mes == 12:  # DICIEMBRE
                        letra_columa = get_column_letter(14)  # COLUMNA N

                    sheet[letra_columa + '8'] = cont_motos  # FILA 8 AF0RO DE MOTOCICLETA
                    sheet[letra_columa + '17'] = cont_autos  # FILA 17 AF0RO DE AUTOS NORMALES
                    sheet[letra_columa + '20'] = cont_autos_1eje  # FILA 20 AF0RO DE AUTOS 1 EJE
                    sheet[letra_columa + '23'] = cont_autos_2eje  # FILA 23 AF0RO DE AUTOS 2 EJE
                    sheet[letra_columa + '33'] = cont_autobuses  # FILA 33 AF0RO AUTOBUSES
                    sheet[letra_columa + '43'] = cont_camion_2ejes + cont_camion_4ejes # FILA 44 AF0RO CAMIONES 2-4 EJES
                    sheet[letra_columa + '47'] = cont_camion_4ejes + cont_camion_5ejes # FILA 47 AF0RO CAMIONES 5-6 EJES
                    sheet[letra_columa + '51'] = cont_camion_7ejes  # FILA 51 AF0RO CAMIONES 7+ EJES
                    mes_anterior = pos_browse.date_order.month
            # Save the spreadsheet
            workbook.save("/tmp/reporte_concentrado_aforo_ingrensos.xlsx")

            # prs.save('/tmp/test.pptx')
            f = open('/tmp/reporte_concentrado_aforo_ingrensos.xlsx', mode="rb")
            return http.request.make_response(f.read(),
                                              [('Content-Type', 'application/octet-stream'),
                                               ('Content-Disposition',
                                                'attachment; filename="{}"'.format('reporte_concentrado_aforo_ingrensos.xlsx'))
                                               ])

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


'''@http.route('/pos/reporte_fin_turno', type='http', auth='user')
def print_sale_details(self, date_start=False, date_stop=False, **kw):
    r = request.env['report.slrc.report_fin_turno']
    pdf, _ = request.env.ref('point_of_sale.reporte_fin_report').with_context(date_start=date_start, date_stop=date_stop).render_qweb_pdf(r)
    pdfhttpheaders = [('Content-Type', 'application/pdf'), ('Content-Length', len(pdf))]
    return request.make_response(pdf, headers=pdfhttpheaders)'''